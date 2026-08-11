# -*- coding: utf-8 -*-
"""
modulos/bodega.py — LA BODEGA: lectura de los archivos de entrada.

Solo lee los Excel de Insumos/ (nunca los modifica) y los convierte
en estructuras de Python listas para que la cocina trabaje con ellas.
"""

import openpyxl

# Ladrillo: VARIABLES constantes con las rutas de los archivos de la Bodega.
RUTA_MAESTRO = 'Insumos/Maestro_Estudiantes.xlsx'
RUTA_REGISTRO = 'Insumos/Registro_Evaluaciones.xlsx'


def cargar_estudiantes(ruta=RUTA_MAESTRO):
    """FUNCIÓN: lee el maestro de estudiantes.

    TIPO de retorno: lista de diccionarios, uno por estudiante.
    Nota: la fila 1 está congelada (encabezados), los datos inician en la fila 2.
    """
    libro = openpyxl.load_workbook(ruta, data_only=True)
    hoja = libro['Estudiantes']
    estudiantes = []
    # Ladrillo: BUCLE for que recorre las filas de datos.
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        # Ladrillo: CONDICIONAL para ignorar filas vacías.
        if fila[0] is None:
            continue
        # Ladrillo: VARIABLE diccionario con los datos de un estudiante.
        estudiantes.append({
            'identificacion': str(fila[0]).strip(),
            'nombre': str(fila[1]).strip(),
            'correo': str(fila[2]).strip(),
            'programa': str(fila[3]).strip(),
            'cohorte': str(fila[4]).strip(),
        })
    return estudiantes


def cargar_evaluaciones(ruta=RUTA_REGISTRO):
    """FUNCIÓN: lee el registro de evaluaciones.

    TIPO de retorno: lista de diccionarios, una entrada por módulo cursado.
    """
    libro = openpyxl.load_workbook(ruta, data_only=True)
    hoja = libro['Evaluaciones']
    evaluaciones = []
    # Ladrillo: BUCLE for que recorre las filas de datos.
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        # Ladrillo: CONDICIONAL para ignorar filas vacías.
        if fila[0] is None:
            continue
        # Ladrillo: VARIABLE diccionario con los datos de una evaluación.
        evaluaciones.append({
            'identificacion': str(fila[0]).strip(),
            'programa': str(fila[1]).strip(),
            'modulo': str(fila[2]).strip(),
            'nota': float(fila[3]),
            'asistencia': float(fila[4]),
            'fecha': str(fila[5]).strip(),
        })
    return evaluaciones
